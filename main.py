import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from tkcalendar import DateEntry  # <--- New import
import mysql.connector
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font

# -------------------- MySQL Connection --------------------
db = mysql.connector.connect(
    host="localhost",
    user="root",
    password="************",  # write your mysql password
    database="customer_mgmt"
)
cursor = db.cursor()

BG_COLOR = "#f5f6fa"
FORM_BG = "#dff9fb"
BTN_COLOR = "#130f40"
BTN_FG = "white"

# -------------------- Login Window --------------------
def show_login_window():
    login_window = tk.Tk()
    login_window.title("Login")
    login_window.geometry("300x200")
    login_window.configure(bg=BG_COLOR)

    tk.Label(login_window, text="Username:", bg=BG_COLOR).pack(pady=5)
    username_entry = tk.Entry(login_window)
    username_entry.pack()
    username_entry.focus()

    tk.Label(login_window, text="Password:", bg=BG_COLOR).pack(pady=5)
    password_entry = tk.Entry(login_window, show="*")
    password_entry.pack()
    
    def check_login():
        entered_user = username_entry.get()
        entered_pass = password_entry.get()
        cursor.execute("SELECT username, role FROM users WHERE username = %s AND password = %s", (entered_user, entered_pass))
        user = cursor.fetchone()

        if user:
            current_user, role = user
            login_window.destroy()
            show_main_window(current_user, role)
        else:
            messagebox.showerror("Login Failed", "Invalid credentials")

    tk.Button(login_window, text="Login", bg=BTN_COLOR, fg=BTN_FG, width=15, command=check_login).pack(pady=20)
    login_window.mainloop()

# -------------------- Main CMS GUI --------------------
def show_main_window(current_user, role):
    def add_customer():
        name = name_entry.get()
        phone = phone_entry.get()
        email = email_entry.get()
        address = address_entry.get()
        category = category_combobox.get()
        tally_serial = tally_serial_entry.get()
        from_date = from_date_entry.get()
        end = end_date_entry.get()
        refer_by = refer_by_combobox.get()
        firm_name = firm_name_entry.get()
        remarks = remark_entry.get()

        if name == "":
            messagebox.showwarning("Input Error", "Name is required")
            return
        if phone == "":
            messagebox.showwarning("Input Error", "Phone number is required")
            return

        sql = """
            INSERT INTO customers 
            (name, phone, email, address, from_date, end_date, added_by, category, tally_serial, refer_by, firm_name, remarks) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        val = (name, phone, email, address, from_date, end, current_user, category, tally_serial, refer_by, firm_name, remarks)
        cursor.execute(sql, val)
        db.commit()

        messagebox.showinfo("Success", "Customer added successfully")
        clear_fields()
        if role == "admin":
            view_customers()

    def clear_fields():
        name_entry.delete(0, tk.END)
        phone_entry.delete(0, tk.END)
        email_entry.delete(0, tk.END)
        address_entry.delete(0, tk.END)
        category_combobox.set("")
        tally_serial_entry.delete(0, tk.END)
        # do not use this : from_date_entry.set_date("")
        from_date_entry.delete(0, tk.END)
        end_date_entry.delete(0, tk.END)
        refer_by_combobox.set("")
        firm_name_entry.delete(0, tk.END)
        remark_entry.delete(0, tk.END)


    def view_customers():
        for item in tree.get_children():
            tree.delete(item)
        cursor.execute("SELECT * FROM customers")
        rows = cursor.fetchall()
        
        for idx, row in enumerate(rows, start=1):
            row_list = [idx, row[1], row[12], row[2], row[3], row[4], row[10], row[9], row[6], row[7], row[8], row[11], row[13]]
            tree.insert("", tk.END, iid=str(row[0]), values=row_list)
    
    def notify_expiring_customers():
        today = datetime.today().date()
        upcoming = today + timedelta(days=5)
        query = "SELECT name, end_date FROM customers WHERE end_date BETWEEN %s AND %s"
        cursor.execute(query, (today, upcoming))
        expiring_customers = cursor.fetchall()

        if expiring_customers:
            message = "Customers nearing end date (within 5 days):\n"
            for name, end_date in expiring_customers:
                message += f"• {name} - ends on {end_date}\n"
            messagebox.showwarning("Upcoming Expiries", message)

    def search_customers():
        keyword = search_entry.get()
        if keyword == "":
            messagebox.showwarning("Search Error", "Enter name or phone to search.")
            return
        for item in tree.get_children():
            tree.delete(item)
        query = "SELECT * FROM customers WHERE LOWER(name) = LOWER(%s) OR phone = %s"
        cursor.execute(query, (keyword, keyword))

        rows = cursor.fetchall()
        for idx, row in enumerate(rows, start=1):
            row_list = [idx, row[1], row[12], row[2], row[3], row[4], row[10], row[9], row[6], row[7], row[8], row[11], row[13]]
            tree.insert("", tk.END, iid=str(row[0]), values=row_list)

    def delete_customer():
        selected = tree.focus()
        if not selected:
            messagebox.showwarning("Selection Error", "Select a customer to delete.")
            return

        customer_id = selected
        confirm = messagebox.askyesno("Confirm Delete", f"Delete customer?")
        if confirm:
            cursor.execute("DELETE FROM customers WHERE id = %s", (customer_id,))
            db.commit()
            view_customers()
            messagebox.showinfo("Deleted", "Customer deleted.")

    def edit_customer():
        selected = tree.focus()
        if not selected:
            messagebox.showwarning("Selection Error", "Select a customer to edit.")
            return
        name = name_entry.get()
        phone = phone_entry.get()
        email = email_entry.get()
        address = address_entry.get()
        category = category_combobox.get()
        tally_serial = tally_serial_entry.get()
        from_date = from_date_entry.get()
        end = end_date_entry.get()
        refer_by = refer_by_combobox.get()
        firm_name = firm_name_entry.get()
        remarks = remark_entry.get()

        if name == "":
            messagebox.showwarning("Input Error", "Name is required")
            return
        sql = """
            UPDATE customers 
            SET name=%s, phone=%s, email=%s, address=%s, category=%s, tally_serial=%s, from_date=%s, end_date=%s ,refer_by=%s, firm_name=%s, remarks=%s
            WHERE id=%s
        """
        val = (name, phone, email, address, category, tally_serial, from_date, end, refer_by, firm_name, remarks, selected)
        cursor.execute(sql, val)
        db.commit()
        messagebox.showinfo("Updated", "Customer updated successfully.")
        if role == "admin":
            view_customers()
        clear_fields()
    
    def filter_by_refer():
        refer_window = tk.Toplevel()
        refer_window.title("Filter by Referred By")
        refer_window.geometry("300x150")

        tk.Label(refer_window, text="Select Refer By Name:", font=("Helvetica", 12)).pack(pady=10)

    # Manually define refer_by options or fetch from DB if needed
        refer_names = ["Kishore Sir", "Padma Mam", "Mahesh", "Rekha", "Suma Sri"]

        selected_refer = tk.StringVar()
        refer_dropdown = ttk.Combobox(refer_window, values=refer_names, textvariable=selected_refer, state="readonly")
        refer_dropdown.pack(pady=10)
        refer_dropdown.set("Choose")

        def fetch_customers_by_refer():
            refer_by = selected_refer.get()
            if refer_by == "Choose" or not refer_by:
                messagebox.showwarning("No Selection", "Please select a name.")
                return

        # Clear current table
            for item in tree.get_children():
                tree.delete(item)

        # Query to filter by refer_by
            cursor.execute("SELECT * FROM customers WHERE refer_by = %s", (refer_by,))
            rows = cursor.fetchall()

            for idx, row in enumerate(rows, start=1):
                row_list = [idx, row[1], row[12], row[2], row[3], row[4], row[10], row[9], row[6], row[7], row[8], row[11], row[13]]
                tree.insert("", tk.END, iid=str(row[0]), values=row_list)

            refer_window.destroy()

        tk.Button(refer_window, text="Filter", command=fetch_customers_by_refer, bg="#2ecc71", fg="white").pack(pady=10)


    def show_renewals():
        renew_window = tk.Toplevel()
        renew_window.title("Customers due for Renewal")

        tree = ttk.Treeview(renew_window, columns=("ID", "Name", "Phone", "End Date"), show="headings")
        for col in ("ID", "Name", "Phone", "End Date"):
            tree.heading(col, text=col)
        tree.pack(fill="both", expand=True)

        def refresh_renewal_list():
            tree.delete(*tree.get_children())
            today = datetime.today().date()
            upcoming = today + timedelta(days=5)
            cursor.execute("SELECT id, name, phone, end_date FROM customers WHERE end_date BETWEEN %s AND %s", (today, upcoming))
            for row in cursor.fetchall():
                tree.insert("", "end", values=row)

        def on_row_click(event):
            selected = tree.focus()
            if not selected:
                return
            values = tree.item(selected, "values")
            customer_id, name, phone, end_date = values
            response = messagebox.askyesno("Renew Subscription", f"Renew {name}'s subscription for 1 more year?")
            if response:
                new_end_date = (datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=365)).date()
                cursor.execute("UPDATE customers SET end_date = %s WHERE id = %s", (new_end_date, customer_id))
                db.commit()
                messagebox.showinfo("Success", f"{name}'s subscription extended to {new_end_date}")
                refresh_renewal_list()

        tree.bind("<Double-1>", on_row_click)
        refresh_renewal_list()
    def show_calls_window():
        calls_window = tk.Toplevel()
        calls_window.title("Call Logs")
        calls_window.geometry("900x500")

        # --- Input Frame ---
        input_frame = tk.Frame(calls_window)
        input_frame.pack(pady=10)

        tk.Label(input_frame, text="Name").grid(row=0, column=0, padx=5)
        name_entry = tk.Entry(input_frame, width=20)
        name_entry.grid(row=0, column=1, padx=5)

        tk.Label(input_frame, text="Phone").grid(row=0, column=2, padx=5)
        phone_entry = tk.Entry(input_frame, width=20)
        phone_entry.grid(row=0, column=3, padx=5)

        tk.Label(input_frame, text="Customer Problem").grid(row=1, column=0, padx=5)
        issue_entry = tk.Entry(input_frame, width=60)
        issue_entry.grid(row=1, column=1, columnspan=3, padx=5, pady=5)

        tk.Label(input_frame, text="Refer By").grid(row=0, column=4, padx=5)
        refer_entry = ttk.Combobox(input_frame, values=["Kishore Sir", "Padma Mam", "mahesh", "rekha", "suma sri"], width=18)
        refer_entry.grid(row=0, column=5)

        def add_call():
            name = name_entry.get()
            phone = phone_entry.get()
            issue = issue_entry.get()
            refer_by = refer_entry.get()
            now = datetime.now()

            if not phone or not name:
                messagebox.showwarning("Missing", "Name and Phone number required.")
                return

            cursor.execute(
                "INSERT INTO calls (name, phone, issue, refer_by, date_logged, status) VALUES (%s, %s, %s, %s, %s, %s)",
                (name, phone, issue, refer_by, now.strftime("%Y-%m-%d %H:%M:%S"), "open")
            )
            db.commit()
            refresh_call_list()

            name_entry.delete(0, tk.END)
            phone_entry.delete(0, tk.END)
            issue_entry.delete(0, tk.END)
            refer_entry.set("")

        def refresh_call_list():
            for item in call_tree.get_children():
                call_tree.delete(item)
            cursor.execute("SELECT id, name, phone, issue, refer_by, date_logged FROM calls WHERE status = 'open'")
            for row in cursor.fetchall():
                call_tree.insert("", tk.END, values=row)

        def mark_resolved():
            selected = call_tree.focus()
            if not selected:
                messagebox.showwarning("Select", "Select an issue to mark done.")
                return

            item_values = call_tree.item(selected)['values']
            call_id = item_values[0]
            
            # Popup for closed_by
            popup = tk.Toplevel()
            popup.title("Closed By")
            popup.geometry("300x150")
            tk.Label(popup, text="Closed By :").pack(pady=10)

            # Determine allowed names based on role
            if role == "admin":
                cursor.execute("SELECT username FROM users")
                closed_by_options = [row[0] for row in cursor.fetchall()]
            else:
                closed_by_options = [current_user]  # only allow their own name

            closed_by_entry = ttk.Combobox(popup, values=closed_by_options, state="readonly")
            closed_by_entry.pack(pady=10)
            closed_by_entry.set(closed_by_options[0])  # auto-select their name

            def confirm_closed_by():
                closed_by = closed_by_entry.get()
                closed_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                # Insert into 'calls' table
                cursor.execute(
                    "UPDATE calls SET status = %s, closed_by = %s, closed_time = %s WHERE id = %s",
                    ("closed", closed_by, closed_time, call_id)
                )
                db.commit()

                popup.destroy()
                refresh_call_list()

            tk.Button(popup, text="Confirm", command=confirm_closed_by, bg="#81ecec").pack(pady=10)

        def show_closed_calls():
            closed_window = tk.Toplevel()
            closed_window.title("Closed Calls")
            closed_window.geometry("1000x400")

            columns = ("ID", "Name", "Phone", "Issue", "Refer By", "Closed By", "Entered Time", "Closed Time")
            closed_tree = ttk.Treeview(closed_window, columns=columns, show="headings")
            for col in columns:
                closed_tree.heading(col, text=col)
                closed_tree.column(col, width=120)
            closed_tree.pack(fill="both", expand=True)

            cursor.execute("SELECT id, name, phone, issue, refer_by, closed_by, date_logged, closed_time FROM calls WHERE status = 'closed'")
            for row in cursor.fetchall():
                closed_tree.insert("", tk.END, values=row)

        def edit_selected_call():
            selected = call_tree.focus()
            if not selected:
                messagebox.showwarning("Select", "Select a call to edit.")
                return
            values = call_tree.item(selected)['values']
            name_entry.delete(0, tk.END)
            name_entry.insert(0, values[1])
            phone_entry.delete(0, tk.END)
            phone_entry.insert(0, values[2])
            issue_entry.delete(0, tk.END)
            issue_entry.insert(0, values[3])
            refer_entry.set(values[4])

            def save_changes():
                updated_values = (
                    name_entry.get(),
                    phone_entry.get(),
                    issue_entry.get(),
                    refer_entry.get(),
                    values[0]  # call_id
                )
                cursor.execute("UPDATE calls SET name=%s, phone=%s, issue=%s, refer_by=%s WHERE id=%s", updated_values)
                db.commit()
                refresh_call_list()
                add_btn.config(text="Add", command=add_call)

            add_btn.config(text="Save", command=save_changes)

        # --- Buttons ---
        btn_frame = tk.Frame(calls_window)
        btn_frame.pack(pady=10)
        add_btn = tk.Button(btn_frame, text="Add", command=add_call, bg="#74b9ff", fg="white", width=15)
        add_btn.grid(row=0, column=0, padx=5)
        if role=="admin":
            tk.Button(btn_frame, text="Edit", command=edit_selected_call, bg="#ffeaa7", width=15).grid(row=0, column=3, padx=5)
        tk.Button(btn_frame, text="Mark Closed (Yes)", command=mark_resolved, bg="#55efc4", width=20).grid(row=0, column=1, padx=5)
        tk.Button(btn_frame, text="View Closed Calls", command=show_closed_calls, bg="#dfe6e9", width=20).grid(row=0, column=2, padx=5)

        # --- Call Table ---
        columns = ("ID", "Name", "Phone", "Issue", "Refer By", "Entered Time")
        call_tree = ttk.Treeview(calls_window, columns=columns, show="headings")
        style = ttk.Style()
        style.configure("Treeview.Heading", anchor="w")  # 'w' means west (left)

        for col in columns:
            call_tree.heading(col, text=col)
            call_tree.column(col, width=120)
        call_tree.pack(fill="both", expand=True, pady=10)

        refresh_call_list()

    def show_monthly_customers_dropdown():
        month_window = tk.Toplevel()
        month_window.title("Select Month")
        month_window.geometry("300x150")
        month_window.resizable(False, False)

        tk.Label(month_window, text="Select a Month", font=("Helvetica", 12)).pack(pady=10)

        months = [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ]

        selected_month = tk.StringVar()
        month_dropdown = ttk.Combobox(month_window, values=months, textvariable=selected_month, state="readonly", width=25)
        month_dropdown.pack(pady=10)
        month_dropdown.set("Choose Month")

        def fetch_customers():
            month = selected_month.get()
            if not month or month == "Choose Month":
                messagebox.showwarning("No Selection", "Please select a month.")
                return

            month_number = months.index(month) + 1

            for item in tree.get_children():
                tree.delete(item)

            cursor.execute("SELECT * FROM customers WHERE MONTH(end_date) = %s", (month_number,))
            rows = cursor.fetchall()

            for idx, row in enumerate(rows, start=1):
                row_list = [idx, row[1], row[2], row[3], row[4], row[10], row[9], row[6], row[7], row[8]]
                tree.insert("", tk.END, iid=str(row[0]), values=row_list)

            month_window.destroy()

        tk.Button(month_window, text="Show Customers", command=fetch_customers, bg="#3498db", fg="white", font=("Helvetica", 10, "bold")).pack(pady=10)
    
    def export_to_excel():
        if not tree.get_children():
            messagebox.showwarning("No Data", "No data to export.")
            return

        file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"customers_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        )
        if not file:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"

        # Write headings
        for col_index, col_name in enumerate(tree["columns"], start=1):
            cell = ws.cell(row=1, column=col_index, value=col_name)
            cell.font = Font(bold=True)

    # Write rows
        for row_index, item_id in enumerate(tree.get_children(), start=2):
            values = tree.item(item_id)["values"]
            for col_index, value in enumerate(values, start=1):
                ws.cell(row=row_index, column=col_index, value=value)

        wb.save(file)
        messagebox.showinfo("Success", f"Data exported to:\n{file}")
    notify_expiring_customers()
    
    root = tk.Tk()
    root.title(f"Customer Management - Logged in as {current_user} ({role})")
    root.geometry("1100x700")

    form_frame = tk.Frame(root)
    form_frame.pack(pady=10)

    tk.Label(form_frame, text="Name").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    name_entry = tk.Entry(form_frame, width=40)
    name_entry.grid(row=0, column=1)
    name_entry.focus()

    tk.Label(form_frame, text="firm Name").grid(row=0, column=2, padx=5, pady=5, sticky="e")
    firm_name_entry = tk.Entry(form_frame, width=40)
    firm_name_entry.grid(row=0, column=3)
    firm_name_entry.focus()

    tk.Label(form_frame, text="Phone").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    phone_entry = tk.Entry(form_frame, width=40)
    phone_entry.grid(row=1, column=1)

    tk.Label(form_frame, text="Email").grid(row=1, column=2, padx=5, pady=5, sticky="e")
    email_entry = tk.Entry(form_frame, width=40)
    email_entry.grid(row=1, column=3)

    tk.Label(form_frame, text="Address").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    address_entry = tk.Entry(form_frame, width=40)
    address_entry.grid(row=2, column=1)

    tk.Label(form_frame, text="Category").grid(row=2, column=2, padx=5, pady=5, sticky="e")
    category_combobox = ttk.Combobox(form_frame, values=["cloud customer", "mitra customer", "1 year customer", "free customer", "One-time customer"], width=37)
    category_combobox.grid(row=2, column=3)

    tk.Label(form_frame, text="Tally Serial Number").grid(row=3, column=0, padx=5, pady=5, sticky="e")
    tally_serial_entry = tk.Entry(form_frame, width=40)
    tally_serial_entry.grid(row=3, column=1)

    tk.Label(form_frame, text="From Date").grid(row=3, column=2, padx=5, pady=5, sticky="e")
    from_date_entry = DateEntry(form_frame, width=37, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
    from_date_entry.grid(row=3, column=3)

    tk.Label(form_frame, text="End Date").grid(row=4, column=0, padx=5, pady=5, sticky="e")
    end_date_entry = DateEntry(form_frame, width=37, background='pink', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
    end_date_entry.grid(row=4, column=1)

    tk.Label(form_frame, text="Referred By").grid(row=4, column=2, padx=5, pady=5, sticky="e")
    # Fetch employee names from users table
    # cursor.execute("SELECT username FROM users WHERE role = 'employee'")
    # employee_names = [row[0] for row in cursor.fetchall()]
    employee_names = ["Kishore Sir", "Padma Mam", "Mahesh", "Rekha", "Suma Sri", "Other"]
    refer_by_combobox = ttk.Combobox(form_frame, values=employee_names, width=37, state="readonly")
    refer_by_combobox.grid(row=4, column=3)

    tk.Label(form_frame, text="remark").grid(row=5, column=0, padx=5, pady=5, sticky="e")
    remark_entry = tk.Entry(form_frame, width=40)
    remark_entry.grid(row=5, column=1)
    remark_entry.focus()

    btn_frame = tk.Frame(root)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="Add Customer", command=add_customer).grid(row=0, column=0, padx=10)
    tk.Button(btn_frame, text="Clear Fields", command=clear_fields).grid(row=0, column=1, padx=10)
    view_btn = tk.Button(btn_frame, text="View All", command=view_customers)
    view_btn.grid(row=0, column=2, padx=10)
    delete_btn = tk.Button(btn_frame, text="Delete Selected", command=delete_customer)
    delete_btn.grid(row=0, column=3, padx=10)
    tk.Button(btn_frame, text="Edit Selected", command=edit_customer).grid(row=0, column=4, padx=10)
    tk.Button(btn_frame, text="Calls", command=show_calls_window).grid(row=0, column=5, padx=10)
    
    search_frame = tk.Frame(root)
    search_frame.pack(pady=10)
    tk.Label(search_frame, text="Search (Name or Phone): ").pack(side=tk.LEFT)
    search_entry = tk.Entry(search_frame, width=30)
    search_entry.pack(side=tk.LEFT, padx=5)
    tk.Button(search_frame, text="Search", command=search_customers).pack(side=tk.LEFT)
    if role=="admin":
        btn2_frame = tk.Frame(root)
        btn2_frame.pack(pady=10)
    
        tk.Button(btn2_frame, text="Renewals", command=show_renewals).grid(row=0, column=0, padx=10)
        # renew_button.pack(pady=10)

        tk.Button(btn2_frame, text="Monthly Customers", command=show_monthly_customers_dropdown).grid(row=0, column=1, padx=10)
        #monthly_btn.pack(pady=5)
    
        tk.Button(btn2_frame, text="Export to Excel", command=export_to_excel).grid(row=0, column=2, padx=10)
        tk.Button(btn2_frame, text="Filter by Refer By", command=filter_by_refer).grid(row=0, column=3, padx=10)

        
    table_frame = tk.Frame(root)
    table_frame.pack(pady=20, fill=tk.BOTH, expand=True)

    # Create vertical and horizontal scrollbars
    tree_scroll_y = tk.Scrollbar(table_frame, orient=tk.VERTICAL)
    tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

    tree_scroll_x = tk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
    tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)

    columns = ("ID", "Name", "firm name", "Phone", "Email", "Address", "FromDate", "EndDate", "AddedBy", "Category", "TallySerial", "refer by", "remarks")

    # Create the Treeview with both scrollbars
    tree = ttk.Treeview(
        table_frame,
        columns=columns,
        show="headings",
        yscrollcommand=tree_scroll_y.set,
        xscrollcommand=tree_scroll_x.set
    )

    # Configure scrollbars to work with the Treeview
    tree_scroll_y.config(command=tree.yview)
    tree_scroll_x.config(command=tree.xview)

    # Define column headers and default widths
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=120, anchor=tk.W)

    tree.pack(fill=tk.BOTH, expand=True)

    
    if role == "employee":
        view_btn.config(state=tk.DISABLED)
        delete_btn.config(state=tk.DISABLED)

    if role == "admin":
        view_customers()

    root.mainloop()

# -------------------- Start App --------------------
show_login_window()
